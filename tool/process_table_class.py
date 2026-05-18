import os
from openpyxl import Workbook, load_workbook
import pandas as pd

from tqdm import tqdm
from urllib.parse import urlparse
import ast
import json
import sys

# 处理导入路径
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

# 尝试从当前目录导入,如果失败则从项目根目录导入
try:
    from json_process import read_json
except ImportError:
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(current_dir)))
    if project_root not in sys.path:
        sys.path.append(project_root)
    from json_process import read_json

def clean_data_for_json(data):
    """清洗数据以便转换为JSON格式"""
    if isinstance(data, dict):
        return {k: clean_data_for_json(v) for k, v in data.items() if pd.notna(v)}
    elif isinstance(data, list):
        return [clean_data_for_json(item) for item in data if pd.notna(item)]
    elif pd.isna(data):
        return None
    else:
        return data


def detect_xml_incompatible_chars(s):
    """
    检测字符串中的XML不兼容字符

    参数:
        s: 输入字符串

    返回:
        list: 包含所有XML不兼容字符信息的列表,每个元素为字典:
              {'position': int, 'char_code': int, 'char_name': str, 'context': str}
    """
    if not isinstance(s, str):
        return []

    issues = []
    # XML不允许的控制字符名称映射
    char_names = {
        0x00: 'NULL',
        0x01: 'SOH', 0x02: 'STX', 0x03: 'ETX',
        0x07: 'BEL', 0x08: 'BS',
        0x0B: 'VT (Vertical Tab)',
        0x0C: 'FF (Form Feed)',
        0x7F: 'DEL'
    }

    for i, char in enumerate(s):
        code = ord(char)
        # XML不允许的控制字符: 0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F, 0x7F
        if (0x00 <= code <= 0x08) or code in [0x0B, 0x0C] or (0x0E <= code <= 0x1F) or code == 0x7F:
            # 获取上下文(前后各20个字符)
            start = max(0, i - 20)
            end = min(len(s), i + 20)
            context = s[start:end]

            issues.append({
                'position': i,
                'char_code': code,
                'char_name': char_names.get(code, f'Control Char'),
                'context': context,
                'context_position': i - start
            })

    return issues


def clean_string_for_xml(s, replacement=' '):
    """
    清洗字符串,移除XML不兼容的控制字符

    参数:
        s: 输入字符串
        replacement: 替换字符,默认为空格

    返回:
        str: 清洗后的字符串
    """
    if not isinstance(s, str):
        return s

    import re
    # 移除XML不允许的控制字符(保留\t, \n, \r)
    # 0x00-0x08, 0x0b, 0x0c, 0x0e-0x1f
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', replacement, s)
    return cleaned
    
def _is_na(value):
    """判断值是否为缺失值，兼容列表/数组等类型（pd.isna对数组返回布尔数组，无法直接用作布尔判断）"""
    if isinstance(value, (list, tuple, dict)):
        return False
    try:
        result = pd.isna(value)
        return bool(result)
    except (ValueError, TypeError):
        return False


class ExcelProcessor:
    def __init__(self, file_path=None, save_path=None, sheet_name=None):
        self.file_path = file_path
        self.save_path = save_path
        self.sheet_name = sheet_name
        self.df = None
        
        if file_path:
            if file_path.endswith('.xlsx'):
                if sheet_name:
                    self.df = pd.read_excel(file_path, sheet_name=sheet_name)
                else:
                    self.df = pd.read_excel(file_path)
    
    def merge_excel_by_name(self, file_b, refer_name='idx', key_name='check_device'):
        """根据name列合并两个Excel文件"""
        df_b = pd.read_excel(file_b, dtype=str)
        
        if key_name in self.df.columns:
            temp_col = f'{key_name}_new'
            merged_df = pd.merge(self.df, df_b[[refer_name, key_name]].rename(columns={key_name: temp_col}), 
                               on=refer_name, how='left')
            mask = merged_df[temp_col].notna()
            merged_df.loc[mask, key_name] = merged_df.loc[mask, temp_col]
            merged_df = merged_df.drop(temp_col, axis=1)
        else:
            merged_df = pd.merge(self.df, df_b[[refer_name, key_name]], on=refer_name, how='left')
        
        # 统计合并结果
        matched_count = merged_df[key_name].notna().sum()
        print(f"成功匹配的行数: {matched_count}")
        print(f"未匹配的行数: {len(merged_df) - matched_count}")

        print(f"原始文件A的行数: {len(self.df)}")
        print(f"原始文件B的行数: {len(df_b)}")
        print(f"合并后的行数: {len(merged_df)}")
        self._save_result(merged_df)
    
    def combine_excel_files(self, file_b):
        # 读取两个 Excel 文件
        df1 = pd.read_excel(self.file_path)
        df2 = pd.read_excel(file_b)
        # 合并数据
        merged_df = pd.concat([df1, df2], ignore_index=True)
        self._save_result(merged_df)

    def split_excel_from_dir(self, dir_path, key_name='gt'):
        """根据文件名 idx_xxx.xxx 分割 Excel 对应列"""
        file_info = []
        file_lists = os.listdir(dir_path)
        files = sorted(file_lists, key=self._sort_key)
        
        for filename in files:
            if filename.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                name, _ = os.path.splitext(filename)
                idx = int(name.split('_')[0])
                file_info.append({'filename': filename, 'idx': idx})
        
        results = []
        for info in file_info:
            idx = info['idx']
            result = self.df[key_name][idx - 2]
            results.append({'name': info['filename'], key_name: result})
        
        self._save_result(pd.DataFrame(results))
    
    def generate_excel_from_dir(self, dir_path, key_name='gt'):
        """从文件目录生成Excel"""
        
        wb = Workbook()
        ws = wb.active
        ws.append(["name", key_name])
        
        file_lists = os.listdir(dir_path)
        # files = sorted(file_lists, key=self._sort_key)
        files = sorted(file_lists)
        
        for file in files:
            if '.DS_Store' in file:
                continue
            else:
                base_name = os.path.splitext(file)[0]
                ws.append([base_name, 0])
        
        wb.save(self.save_path)
        print(f"已从{dir_path}生成Excel并保存到: {self.save_path}")
    
    def extract_rows_by_file_names(self, dir_path):
        """根据文件名称提取行"""
        file_names = []
        for filename in os.listdir(dir_path):
            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
                base_name = os.path.splitext(filename)[0]
                base_name = base_name.split('_')[-1]
                file_names.append(base_name)
        
        filtered_df = self.df[self.df['name'].isin(file_names)]
        self._save_result(filtered_df)
    
    def update_excel_with_labels(self, label_dir, suffix='.txt'):
        """更新Excel标签"""
        wb = load_workbook(self.file_path)
        ws = wb.active
        label_files = os.listdir(label_dir) # 遍历标签目录中的文件
        
        # 遍历Excel表格中的行（跳过表头）
        for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            file_name = row_values[0] + suffix
            if file_name in label_files:
                ws.cell(row=row_idx, column=2, value=1) # 假设cell_phone列是第二列
        
        wb.save(self.save_path)
    
    def get_specify_keyvalue_from_table(self, key_name):
        """从原始 url 表格获取指定键值"""
        wb = Workbook()
        ws = wb.active
        ws.append(["name", key_name])
        
        for idx, url in enumerate(tqdm(self.df.source_url)):
            parsed_url = urlparse(url)
            filename = os.path.basename(parsed_url.path)
            filename = filename.split('.')[0]
            filename = str(idx + 2) + '_' + filename
            key_value = self.df[key_name][idx]
            ws.append([filename, key_value])
        
        wb.save(self.save_path)
        print(f"已根据{key_name}提取数据并保存到: {self.save_path}")
    
    def compare_columns(self, col1, col2, diff_col):
        """比较列"""
        self.df[diff_col] = 0
        self.df.loc[(self.df[col1] != self.df[col2]), diff_col] = 1
        self._save_result(self.df)
    
    def judge_columns(self, col1, col2, diff_col):
        """判断列"""
        self.df[diff_col] = 0
        for idx, row in self.df.iterrows():
            if row[col1] >= 2 or row[col2] >= 2:
                self.df.at[idx, diff_col] = 1
        self._save_result(self.df)
    
    def group_by_split_name(self, split_name):
        """按名称分组"""
        if split_name not in self.df.columns:
            print(f"错误：Excel文件中不存在'{split_name}'列")
            return
        
        with pd.ExcelWriter(self.save_path, engine='openpyxl') as writer:
            check_types = self.df[split_name].unique()
            for check_type in check_types:
                filtered_df = self.df[self.df[split_name] == check_type]
                sheet_name = str(check_type)[:31]
                filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"已按{split_name}分组并保存到: {self.save_path}")
        print(f"共发现{len(check_types)}个不同的{split_name}类型")
    
    def _save_result(self, df):
        """保存结果到Excel文件"""
        df.to_excel(self.save_path, index=False)
        print(f"已保存结果到: {self.save_path}")
    

    def mark_existing_ids(self, file_b, id_col='id', has_col='has'):
        """
        标记a文件中id是否存在于b文件中
        
        参数:
        file_b: b文件路径
        id_col: id列名，默认为'id'
        has_col: 新增的标记列名，默认为'has'
        """
        # 1. 读取b文件中的id列表
        df_b = pd.read_excel(file_b)
        b_ids = set(df_b[id_col].unique())  # 使用set提高查找效率
        
        # 2. 在a中添加has列
        self.df[has_col] = self.df[id_col].isin(b_ids).astype(int)
        
        # 3. 统计结果
        total_a = len(self.df)
        matched = self.df[has_col].sum()
        print(f"文件A总行数: {total_a}")
        print(f"在B中找到的id数: {matched}")
        print(f"未找到的id数: {total_a - matched}")
        
        # 4. 保存结果
        self._save_result(self.df)

    def update_column_from_df_b(self, file_b, refer_name='idx', key_name='key_name'):
        """
        从文件B中按idx更新指定列到self.df
        
        参数:
        file_b: 包含更新数据的Excel文件路径
        refer_name: 匹配列名，默认为'idx'
        key_name: 要更新的列名，默认为'key_name'
        """
        try:
            # 1. 读取文件B
            print(f"正在读取文件B: {file_b}")
            df_b = pd.read_excel(file_b)
            
            # 2. 验证必要的列是否存在
            if refer_name not in df_b.columns:
                print(f"错误：文件B中不存在'{refer_name}'列")
                print(f"文件B的列名: {list(df_b.columns)}")
                return
            
            if key_name not in df_b.columns:
                print(f"错误：文件B中不存在'{key_name}'列")
                print(f"文件B的列名: {list(df_b.columns)}")
                return
            
            if refer_name not in self.df.columns:
                print(f"错误：文件A中不存在'{refer_name}'列")
                print(f"文件A的列名: {list(self.df.columns)}")
                return
            
            # 3. 数据类型处理 - 统一转换为字符串类型避免合并错误
            print("正在处理数据类型...")
            df_b_copy = df_b.copy()
            df_a_copy = self.df.copy()
            
            # 确保匹配列为字符串类型
            df_b_copy[refer_name] = df_b_copy[refer_name].astype(str)
            df_a_copy[refer_name] = df_a_copy[refer_name].astype(str)
            
            # 处理NaN值
            df_b_copy[refer_name] = df_b_copy[refer_name].fillna('')
            df_a_copy[refer_name] = df_a_copy[refer_name].fillna('')
            
            # 去除空格
            df_b_copy[refer_name] = df_b_copy[refer_name].str.strip()
            df_a_copy[refer_name] = df_a_copy[refer_name].str.strip()
            
            print(f"文件A总行数: {len(df_a_copy)}")
            print(f"文件B总行数: {len(df_b_copy)}")
            print(f"文件A中'{refer_name}'唯一值数量: {df_a_copy[refer_name].nunique()}")
            print(f"文件B中'{refer_name}'唯一值数量: {df_b_copy[refer_name].nunique()}")
            
            # 4. 执行更新操作
            print(f"正在从文件B更新'{key_name}'列...")
            
            # 使用左连接保留self.df的所有行
            merged_df = pd.merge(
                df_a_copy, 
                df_b_copy[[refer_name, key_name]], 
                on=refer_name, 
                how='left',
                suffixes=('', '_b')
            )
            
            # 如果key_name已存在，直接更新；否则添加新列
            if key_name in self.df.columns:
                # 更新已存在的列
                mask = merged_df[f'{key_name}_b'].notna()
                merged_df.loc[mask, key_name] = merged_df.loc[mask, f'{key_name}_b']
                merged_df = merged_df.drop(f'{key_name}_b', axis=1)
            else:
                # 添加新列
                merged_df = merged_df.rename(columns={f'{key_name}_b': key_name})
            
            # 5. 统计结果
            matched_count = merged_df[key_name].notna().sum()
            total_rows = len(merged_df)
            
            print(f"成功匹配的行数: {matched_count}")
            print(f"未匹配的行数: {total_rows - matched_count}")
            
            # 6. 更新self.df并保存
            self.df = merged_df
            self._save_result(self.df)
            print("更新完成！")
            
        except Exception as e:
            print(f"更新过程中出现错误: {str(e)}")
            print("错误详情:")
            import traceback
            traceback.print_exc()

    def table2json(self, save_path, key_name=None, keep_columns=None):
        """将表格转换为JSON"""
        new_data = []
        if keep_columns is not None:
            available_cols = [col for col in keep_columns if col in self.df.columns]
            self.df = self.df[available_cols]
        if key_name is not None:
            for _, row in self.df.iterrows():
                single_data = ast.literal_eval(str(row[key_name]))
                new_data.append(single_data)
        else:
            data = self.df.to_dict('records')
            new_data = clean_data_for_json(data)

        with open(save_path, 'w', encoding='utf-8') as f:
            json.dump(new_data, f, ensure_ascii=False, indent=2)
        print(f"已保存结果到: {save_path}")
    
    def json2table(self, file_path, save_path, keep_columns=None, auto_clean=True, verbose=True):
        """
        将JSON转换为表格

        参数:
            file_path: JSON文件路径
            save_path: 输出文件路径(.xlsx 或 .csv)
            keep_columns: 保留的列名列表,例如['capture_id','pred'],为None时保留全部列
            auto_clean: 是否自动清洗XML不兼容字符,默认True
            verbose: 是否输出详细日志,默认True
        """
        # 1. 读取JSON数据
        data = read_json(file_path)
        df = pd.DataFrame(data)

        if keep_columns is not None:
            available_cols = [col for col in keep_columns if col in df.columns]
            if verbose:
                skipped = [col for col in keep_columns if col not in df.columns]
                if skipped:
                    print(f"以下列不存在,已跳过: {skipped}")
                print(f"保留列: {available_cols}")
            df = df[available_cols]

        if verbose:
            print(f"读取了 {len(df)} 行数据, {len(df.columns)} 列")

        # 2. 检测XML不兼容字符
        if save_path.endswith('.xlsx'):
            if verbose:
                print("\n正在检测数据中的XML不兼容字符...")

            report = self._detect_dataframe_issues(df)

            if report['has_issues']:
                # 打印检测报告
                self._print_detection_report(report)

            # 自动清洗(无论是否检测到问题,都进行清洗以确保安全)
            if auto_clean:
                if verbose:
                    print("\n正在自动清洗数据...")
                df = self._clean_dataframe(df, replacement=' ', verbose=verbose)
                if verbose:
                    print("清洗完成!\n")

        # 3. 保存文件
        try:
            if save_path.endswith('.xlsx'):
                df.to_excel(save_path, index=False)
            elif save_path.endswith('.csv'):
                df.to_csv(save_path, index=False)
            print(f"已保存结果到: {save_path}")
        except Exception as e:
            print(f"保存文件时出错: {e}")
            print("提示: 如果仍然遇到XML不兼容字符错误,请尝试手动清洗数据")
            raise

    def _detect_dataframe_issues(self, df):
        """
        检测DataFrame中的XML不兼容字符问题

        参数:
            df: pandas DataFrame

        返回:
            dict: 检测报告,包含:
                  - 'has_issues': bool, 是否有问题
                  - 'total_issues': int, 问题总数
                  - 'affected_fields': list, 受影响的字段列表
                  - 'details': list, 问题详情列表
        """
        all_issues = []
        affected_fields = set()

        # 遍历所有列
        for col in df.columns:
            for idx, value in df[col].items():
                # 跳过缺失值（列表/数组类型的pd.isna返回数组，无法直接用作布尔判断）
                if _is_na(value):
                    continue
                # 将值转换为字符串进行检查
                str_value = str(value)
                issues = detect_xml_incompatible_chars(str_value)
                if issues:
                    affected_fields.add(col)
                    for issue in issues:
                        all_issues.append({
                            'row': idx,
                            'field': col,
                            **issue
                        })

        return {
            'has_issues': len(all_issues) > 0,
            'total_issues': len(all_issues),
            'affected_fields': list(affected_fields),
            'details': all_issues
        }

    def _print_detection_report(self, report):
        """
        打印检测报告

        参数:
            report: _detect_dataframe_issues返回的报告字典
        """
        print("\n" + "="*70)
        print("XML不兼容字符检测报告")
        print("="*70)
        print(f"发现 {report['total_issues']} 处XML不兼容字符问题")
        print(f"受影响的字段: {', '.join(report['affected_fields'])}")

        # 显示前10个问题详情
        print("\n问题详情 (显示前10个):")
        for i, issue in enumerate(report['details'][:10]):
            print(f"\n[问题 {i+1}]")
            print(f"  位置: 第 {issue['row']} 行, 字段 '{issue['field']}'")
            print(f"  字符: 0x{issue['char_code']:02X} ({issue['char_name']})")
            print(f"  字符位置: 第 {issue['position']} 个字符")
            print(f"  上下文: ...{repr(issue['context'])}...")
            print(f"  位置标记: {' ' * issue['context_position']}^")

        if len(report['details']) > 10:
            print(f"\n... 还有 {len(report['details']) - 10} 个问题未显示")

        # 统计信息
        from collections import Counter
        field_counter = Counter([issue['field'] for issue in report['details']])
        char_counter = Counter([issue['char_code'] for issue in report['details']])

        print("\n按字段统计:")
        for field, count in field_counter.most_common():
            print(f"  {field}: {count} 处问题")

        print("\n按字符类型统计:")
        char_names = {
            0x00: 'NULL (0x00)',
            0x0B: 'Vertical Tab (0x0B)',
            0x0C: 'Form Feed (0x0C)'
        }
        for char_code, count in char_counter.most_common():
            char_name = char_names.get(char_code, f'0x{char_code:02X}')
            print(f"  {char_name}: {count} 处")

        print("="*70)

    def _clean_dataframe(self, df, replacement=' ', verbose=True):
        """
        清洗DataFrame中的所有字符串列

        参数:
            df: pandas DataFrame
            replacement: 替换字符,默认为空格
            verbose: 是否打印清洗进度

        返回:
            DataFrame: 清洗后的DataFrame
        """
        df_cleaned = df.copy()
        cleaned_count = 0

        for col in df_cleaned.columns:
            # 检查该列是否有问题
            has_issues = False
            for value in df_cleaned[col]:
                if not _is_na(value):
                    str_value = str(value)
                    if detect_xml_incompatible_chars(str_value):
                        has_issues = True
                        break

            if has_issues:
                # 清洗该列 - 对所有非空值转换为字符串后清洗
                def clean_value(x):
                    if not _is_na(x):
                        str_x = str(x)
                        return clean_string_for_xml(str_x, replacement)
                    return x

                df_cleaned[col] = df_cleaned[col].apply(clean_value)
                cleaned_count += 1
                if verbose:
                    print(f"  已清洗列 '{col}'")

        if verbose:
            print(f"共清洗 {cleaned_count} 列")

        return df_cleaned

    @staticmethod
    def _sort_key(filename):
        """文件名排序键"""
        if '.DS' in filename or 'classes' in filename:
            return 100000
        first_char_after_strip = int(filename.split('_')[0])
        return first_char_after_strip

def process_command_line():
    """处理命令行参数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel处理工具')
    
    # 基础参数
    parser.add_argument('-f', '--file_path', type=str, help='输入Excel文件路径')
    parser.add_argument('-s', '--save_path', type=str, help='输出Excel文件路径')
    parser.add_argument('-n', '--sheet_name', type=str, default=None, help='工作表名称')
    
    # 功能选择参数
    parser.add_argument('-m', '--merge', action='store_true', help='合并两个Excel文件')
    parser.add_argument('--combine', action='store_true', help='合并两个Excel文件')
    parser.add_argument('-p', '--split', action='store_true', help='根据图片目录分割Excel')
    parser.add_argument('-g', '--gen', action='store_true', help='从图片目录生成Excel')
    parser.add_argument('-e', '--extract', action='store_true', help='根据图片名称提取行')
    parser.add_argument('-u', '--update', action='store_true', help='更新Excel标签')
    parser.add_argument('-k', '--get_fromkey', action='store_true', help='从表格获取指定键值')
    parser.add_argument('-c', '--compare', action='store_true', help='比较列')
    parser.add_argument('-j', '--judge', action='store_true', help='判断列')
    parser.add_argument('-b', '--group', action='store_true', help='按名称分组')
    parser.add_argument('--mark_ids', action='store_true', help='标记id')
    parser.add_argument('--update_col', action='store_true', help='从文件B更新指定列')
    parser.add_argument('--table2json', action='store_true', help='将表格转换为JSON')
    parser.add_argument('--json2table', action='store_true', help='将JSON转换为表格')

    # 额外参数
    parser.add_argument('-i', '--input_dir', type=str, help='文件目录路径')
    parser.add_argument('--excel_2', type=str, help='第二个Excel文件路径')
    parser.add_argument('--key_name', type=str, default=None, help='键名')
    parser.add_argument('-1', '--col1', type=str, help='第一列名')
    parser.add_argument('-2', '--col2', type=str, help='第二列名')
    parser.add_argument('-d', '--diff_col', type=str, help='差异列名')
    parser.add_argument('--split_name', type=str, help='分组名称')
    parser.add_argument('--keep_columns', type=str, default=None, help='保留的列名,逗号分隔,例如capture_id,pred')

    return parser.parse_args()

def main():
    args = process_command_line()
    
    # 检查必要的参数
    # if args.file_path is None and not args.gen:
    #     print("错误：必须提供输入文件路径 (-f/--file_path)")
    #     exit(1)
    
    if args.save_path is None:
        print("错误：必须提供输出文件路径 (-s/--save_path)")
        exit(1)
    
    # 创建处理器实例
    processor = ExcelProcessor(args.file_path, args.save_path, args.sheet_name)
    
    # 根据不同的功能标志执行相应的操作
    if args.merge:
        if args.excel_2 is None:
            print("错误：合并Excel需要提供第二个Excel文件路径 (-2/--excel_2)")
            exit(1)
        processor.merge_excel_by_name(args.excel_2, key_name=args.key_name)
    
    elif args.combine:
        if args.excel_2 is None:
            print("错误：合并Excel需要提供第二个Excel文件路径 (-2/--excel_2)")
            exit(1)
        processor.combine_excel_files(args.excel_2)

    elif args.split:
        if args.input_dir is None:
            print("错误：分割Excel需要提供文件目录路径 (-i/--input_dir)")
            exit(1)
        processor.split_excel_from_dir(args.input_dir, args.key_name)
    
    elif args.gen:
        if args.input_dir is None:
            print("错误：生成Excel需要提供文件目录路径 (-i/--input_dir)")
            exit(1)
        processor.generate_excel_from_dir(args.input_dir)
    
    elif args.extract:
        if args.input_dir is None:
            print("错误：提取行需要提供文件目录路径 (-i/--input_dir)")
            exit(1)
        processor.extract_rows_by_file_names(args.input_dir)
    
    elif args.update:
        if args.input_dir is None:
            print("错误：更新标签需要提供图片目录路径 (-i/--input_dir)")
            exit(1)
        processor.update_excel_with_labels(args.input_dir)
    
    elif args.get_fromkey:
        processor.get_specify_keyvalue_from_table(args.key_name)
    
    elif args.compare:
        if not all([args.col1, args.col2, args.diff_col]):
            print("错误：比较列需要提供col1、col2和diff_col参数")
            exit(1)
        processor.compare_columns(args.col1, args.col2, args.diff_col)
    
    elif args.judge:
        if not all([args.col1, args.col2, args.diff_col]):
            print("错误：判断列需要提供col1、col2和diff_col参数")
            exit(1)
        processor.judge_columns(args.col1, args.col2, args.diff_col)
    
    elif args.group:
        if args.split_name is None:
            print("错误：分组需要提供split_name参数 (--split_name)")
            exit(1)
        processor.group_by_split_name(args.split_name)
    
    elif args.mark_ids:
        if args.excel_2 is None:
            print("错误：标记id需要提供第二个Excel文件路径 (-2/--excel_2)")
            exit(1)
        processor.mark_existing_ids(args.excel_2, args.key_name)
    
    elif args.update_col:
        if args.excel_2 is None:
            print("错误：更新列需要提供第二个Excel文件路径 (-2/--excel_2)")
            exit(1)
        processor.update_column_from_df_b(args.excel_2, key_name=args.key_name)
    
    elif args.table2json:
        # if args.key_name is None:
        #     keys = [k.strip() for k in args.keep_keys.split(',')]
        keys = None
        processor.table2json(args.save_path, key_name=args.key_name, keep_columns=keys)
    elif args.json2table:
        keep_cols = [k.strip() for k in args.keep_columns.split(',')] if args.keep_columns else None
        processor.json2table(args.file_path, args.save_path, keep_columns=keep_cols)
        
    else:
        print("错误：请指定要执行的功能")
        print("可用的功能：")
        print("  -m/--merge    根据规则，合并两个Excel文件")
        print("  -p/--split    根据图片目录分割Excel")
        print("  -g/--gen      从图片目录生成Excel")
        print("  -e/--extract  根据图片名称提取行")
        print("  -u/--update   更新Excel标签")
        print("  -k/--key      从表格获取指定键值")
        print("  -c/--compare  比较列")
        print("  -j/--judge    判断列")
        print("  -b/--group    按名称分组")
        print("  --update_col   从文件B更新指定列")

if __name__ == '__main__':
    main()